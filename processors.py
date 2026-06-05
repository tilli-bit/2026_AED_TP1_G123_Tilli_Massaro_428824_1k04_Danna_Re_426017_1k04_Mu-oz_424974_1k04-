import os
import logging

logger = logging.getLogger(__name__)


def extract_text(file_path: str, filename: str, content_type: str) -> str | None:
    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext == ".pdf":
            return extract_pdf(file_path)
        elif ext in (".mp3", ".wav", ".m4a", ".webm", ".ogg"):
            return extract_audio(file_path)
        elif ext in (".xlsx", ".xls"):
            return extract_excel(file_path)
        elif ext == ".csv":
            return extract_csv(file_path)
        elif ext in (".txt", ".md"):
            return extract_text_file(file_path)
        elif ext == ".eml":
            return extract_email(file_path)
        else:
            logger.warning(f"Unsupported file type: {ext}")
            return None
    except Exception as e:
        logger.error(f"Error processing {filename}: {e}")
        return None


def extract_pdf(file_path: str) -> str:
    import pdfplumber
    texts = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
    return "\n\n".join(texts)


def extract_audio(file_path: str) -> str:
    import whisper
    model = whisper.load_model("base")
    result = model.transcribe(file_path)
    return result["text"]


def extract_excel(file_path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append("\t".join(cells))
        parts.append(f"Sheet: {sheet}\n" + "\n".join(rows))
    return "\n\n".join(parts)


def extract_csv(file_path: str) -> str:
    import pandas as pd
    df = pd.read_csv(file_path)
    return df.to_string(index=False)


def extract_text_file(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def extract_email(file_path: str) -> str:
    import email
    from email import policy
    with open(file_path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=policy.default)
    parts = []
    parts.append(f"From: {msg.get('From', '')}")
    parts.append(f"To: {msg.get('To', '')}")
    parts.append(f"Subject: {msg.get('Subject', '')}")
    parts.append(f"Date: {msg.get('Date', '')}")
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                body += part.get_content()
    else:
        body = msg.get_content()
    parts.append(f"\n{body}")
    return "\n".join(parts)


def get_file_type(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    type_map = {
        ".pdf": "PDF",
        ".mp3": "Audio",
        ".wav": "Audio",
        ".m4a": "Audio",
        ".webm": "Audio",
        ".ogg": "Audio",
        ".xlsx": "Excel",
        ".xls": "Excel",
        ".csv": "CSV",
        ".txt": "Texto",
        ".md": "Markdown",
        ".eml": "Email",
    }
    return type_map.get(ext, "Otro")
